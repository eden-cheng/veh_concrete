from general_tool import yaml_parser
import re
import openpyxl
import os

case_group_path = "./parser_tool/case_group.yaml"
path = "./parser_result/"
relation = "./parser_tool/relation.yaml"

class Parser():
    def __init__(self, ws, name1, name2, first_column, first_row):
        self.ws = ws
        self.name1 = name1
        self.name2 = name2
        self.first_column = first_column
        self.first_row = first_row
        self.obj = yaml_parser.Parser()
        self.dic = {}
        self.func_relation = {}
    
    def combine(self):
        self.row_parser()
        self.column_parser()

    def func(self, temp, path, name):
        key = self.name1 + "_" + self.name2 + "_" + temp
        self.dic[key] = path + name

    def func_case_row(self, content):
        """
        因为现在excel中做了多表关联，提取出来不是真实内容而是函数
        通过这个方法建立函数名与case id的对应关系
        """
        path_yaml = "./parser_result/" + self.name1 + '_case_row.yaml'
        case_row_dic = self.obj.yaml_manage(path_yaml)
        temp001 = {}
        for id01,id02 in case_row_dic.items():
            for id02_key, id02_value in id02.items():
                temp001[id02_key] = id02_value
        if content:
            content_num = re.findall("\d+", content)
            if content_num:
                for id, row in temp001.items():
                    if int(content_num[0]) == row:
                        self.func_relation[content] = id

    def row_parser(self):
        column_title = self.ws[self.first_column]
        dic = {}
        for cell in column_title[:2000]:
            # print(cell.value)
            dic[cell.value] = cell.row
        if self.name2 == 'case':
            dic = self.arrange(dic).copy()
            # print(dic)
        if self.name2 == 'para':
            temp = {}
            temp01 = {}
            #生成函数与caseid的映射关系
            for key, value in dic.items():
                self.func_case_row(key)
            #将二级标题的函数替换成caseid
            for key, value in dic.items():
                if key in self.func_relation:
                    temp[self.func_relation[key]] = value
                else:
                    temp[key] = value
            #将一级标题的函数替换成caseid
            for key, value in self.arrange(temp).items():
                id_arr = tuple(value.keys())[0].split('_')
                temp01[id_arr[0]+ '_' + id_arr[1]] = value
            dic = temp01.copy()
        # print(dic)
        #获取文件路径
        name = self.name1 + '_' + self.name2 + "_row.yaml"
        self.func('row', path, name)
        self.obj.yaml_generate(dic, path, name)

        # self.obj.yaml_generate(self.arrange(dic), path, name)
                

    def column_parser(self):
        row_title = self.ws[self.first_row]
        dic = {}
        for cell in row_title[:100]:
            # print(cell.value)
            dic[cell.value] = cell.column   #由于实车参数库中 cell.value 的内容都相同，放在字典时会覆盖前面的部分
        name = self.name1 + '_' + self.name2 + "_column.yaml"
        self.func('column', path, name)
        self.obj.yaml_generate(dic, path, name)

    def arrange(self, data):
        dic = {}
        temp = {}
        for key, value in data.items():
            # print(key)
            # print(self.name1)
            if key:
                pattern = re.compile('_')
                result = pattern.findall(key)
                # print(len(result))
                if len(result) == 1:
                    temp = {}
                    id_01 = key.strip()   #增加删除空格功能
                    dic[id_01] = temp
                if len(result) == 2:
                    id_02 = key.strip()  #增加删除空格功能
                    temp[id_02] = value
        return dic

class Excel():
    def __init__(self):
        self.obj = yaml_parser.Parser()

    def import_data(self, wb, file_name, arr, odd, feature):
        yaml_path_file = "file_path.yaml"
        titles_path = "titles.yaml"

        """将output整理的列表，写入表格，仅做写入，不做列表的追加等操作"""
        # print(feature)
        ws_title = feature + '_' + odd
        # print(ws_title)
        ws = wb.create_sheet(ws_title)
        # print(wb.sheetnames)
        titles = self.obj.yaml_manage(titles_path)['output_titles']
        #将标题填入excel 
        for i in range(len(titles)):
            ws.cell(row = 1, column = i + 1).value = titles[i]
        #将expand整理出的case，从字典形式填入Excel中
        j = 2
        for dic in arr:
            k = 1
            # print(dic)
            for title in titles:
                if title in dic.keys():
                    ws.cell(row = j, column = k).value = dic[title]
                k += 1
            j = j + 1
        wb.save(file_name)

    def wb_get(self, file_path):
        return openpyxl.load_workbook(file_path)

    def wb_new(self):
        wb = openpyxl.Workbook()
        ws_default = wb.active
        wb.remove(ws_default)   #移除默认的表单，方便后面创建新的表单
        return wb

    def create_folder(self, file_path):
        isExists = os.path.exists(file_path)
        if not isExists:
            os.makedirs(file_path)
